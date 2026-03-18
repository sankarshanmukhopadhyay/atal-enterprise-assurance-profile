#!/usr/bin/env python3
"""
export_csv_xlsx.py — Export the control catalog or an assessment summary into
stakeholder-friendly CSV and XLSX files.

Usage:
    python scripts/export_csv_xlsx.py --source catalog
    python scripts/export_csv_xlsx.py --source assessment --result assessments/samples/eap-l1-sample-assessment.json
    python scripts/export_csv_xlsx.py --source catalog --format xlsx
"""

import sys
import json
import csv
import pathlib
import argparse


REPO_ROOT = pathlib.Path(__file__).parent.parent
CATALOG_PATH = REPO_ROOT / "catalogs" / "atal-eap-control-catalog.json"
OUTPUT_DIR = REPO_ROOT / "artifacts"

VALID_SOURCES = ["catalog", "assessment"]
VALID_FORMATS = ["csv", "xlsx", "all"]


def load_json(path: pathlib.Path) -> dict:
    return json.loads(path.read_text(encoding="utf-8"))


def catalog_rows(catalog: dict) -> tuple[list[str], list[list]]:
    headers = [
        "Control ID", "Title", "Family", "Assurance Levels", "Normative Strength",
        "Control Type", "Criticality", "Statement", "ATAL References",
        "Repo Sources", "Automation Candidate"
    ]
    rows = []
    for ctrl in catalog["controls"]:
        rows.append([
            ctrl["id"],
            ctrl["title"],
            ctrl["family"],
            ", ".join(ctrl.get("assurance_levels", [])),
            ctrl["normative_strength"],
            ctrl["control_type"],
            ctrl["criticality"],
            ctrl["statement"],
            "; ".join(ctrl.get("mapped_atal_references", [])),
            "; ".join(ctrl.get("mapped_repo_sources", [])),
            "Yes" if ctrl.get("automation_candidate") else "No",
        ])
    return headers, rows


def assessment_rows(result: dict) -> tuple[list[str], list[list]]:
    headers = [
        "Control ID", "Result", "Finding", "Remediation", "Waiver Ref"
    ]
    rows = []
    for item in result.get("control_results", []):
        rows.append([
            item.get("control_id", ""),
            item.get("result", ""),
            item.get("finding", ""),
            item.get("remediation", ""),
            item.get("waiver_ref", ""),
        ])
    return headers, rows


def write_csv(headers: list[str], rows: list[list], filename: str):
    OUTPUT_DIR.mkdir(exist_ok=True)
    out = OUTPUT_DIR / filename
    with open(out, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(headers)
        writer.writerows(rows)
    print(f"  Written: {out}")


def write_xlsx(headers: list[str], rows: list[list], filename: str, sheet_name: str):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        print("ERROR: openpyxl not installed. Run: pip install openpyxl", file=sys.stderr)
        sys.exit(2)

    OUTPUT_DIR.mkdir(exist_ok=True)
    out = OUTPUT_DIR / filename

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="2E4057")
    header_alignment = Alignment(horizontal="center", wrap_text=True)

    ws.append(headers)
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    alt_fill = PatternFill("solid", fgColor="F0F4F8")
    for i, row in enumerate(rows, start=2):
        ws.append(row)
        if i % 2 == 0:
            for cell in ws[i]:
                cell.fill = alt_fill
        for cell in ws[i]:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    # Auto-size columns (approximate)
    for col in ws.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 60)

    ws.freeze_panes = "A2"
    wb.save(out)
    print(f"  Written: {out}")


def main():
    parser = argparse.ArgumentParser(description="Export EAP catalog or assessment summary to CSV/XLSX.")
    parser.add_argument("--source", required=True, choices=VALID_SOURCES, help="What to export")
    parser.add_argument("--result", default=None, help="Path to assessment result JSON (required if --source assessment)")
    parser.add_argument("--format", default="all", choices=VALID_FORMATS)
    args = parser.parse_args()

    if args.source == "assessment" and not args.result:
        print("ERROR: --result is required when --source is 'assessment'", file=sys.stderr)
        sys.exit(2)

    if args.source == "catalog":
        if not CATALOG_PATH.exists():
            print(f"ERROR: Catalog not found: {CATALOG_PATH}", file=sys.stderr)
            sys.exit(2)
        print(f"Loading catalog: {CATALOG_PATH}")
        catalog = load_json(CATALOG_PATH)
        headers, rows = catalog_rows(catalog)
        csv_name = "eap-control-catalog-export.csv"
        xlsx_name = "eap-control-catalog-export.xlsx"
        sheet_name = "EAP Control Catalog"
    else:
        result_path = pathlib.Path(args.result)
        if not result_path.exists():
            print(f"ERROR: File not found: {result_path}", file=sys.stderr)
            sys.exit(2)
        print(f"Loading assessment result: {result_path}")
        result = load_json(result_path)
        headers, rows = assessment_rows(result)
        level = result.get("assurance_level", "EAP")
        assess_id = result.get("assessment_id", "assessment")
        csv_name = f"{assess_id}-export.csv"
        xlsx_name = f"{assess_id}-export.xlsx"
        sheet_name = f"{level} Assessment"

    print(f"  Rows: {len(rows)}")

    fmt = args.format
    if fmt in ("csv", "all"):
        write_csv(headers, rows, csv_name)
    if fmt in ("xlsx", "all"):
        write_xlsx(headers, rows, xlsx_name, sheet_name)

    print("Done.")


if __name__ == "__main__":
    main()
